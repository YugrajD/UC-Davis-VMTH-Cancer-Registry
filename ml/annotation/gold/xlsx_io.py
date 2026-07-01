"""Read/write the Excel review workbook for gold annotation.

The workbook handed to the professional has four sheets:
  - "Review"       one row per diagnosis.  Dropdowns for verdict / confirmed_group
                   / confirmed_term; the cascade suggestion is prefilled (greyed);
                   report context (HIST SUMMARY + FINAL COMMENT) is shown for
                   negation/hedging judgement.  Rows needing a correction light up.
  - "Instructions" how to fill it in.
  - "Taxonomy"     full (Group, Term, Code) reference for lookup.
  - "Lists"        hidden dropdown sources (unique groups, unique terms).

``read_review_rows`` returns the Review sheet as a list of dicts keyed by column
header, so ``ingest`` / ``consistency`` consume it exactly like the old CSV path.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REVIEW_SHEET = "Review"
VERDICT_CHOICES = ["correct", "wrong", "no_cancer", "uncertain"]

# Column groups (by header name) used for styling and validation placement.
_CASCADE_COLS = (
    "cascade_matched_term", "cascade_matched_group",
    "cascade_matched_code", "cascade_method",
)
_HUMAN_COLS = ("verdict", "confirmed_term", "confirmed_group", "notes")
_WRAP_COLS = ("diagnosis", "HISTOPATHOLOGICAL SUMMARY", "FINAL COMMENT")
_WIDTHS = {
    "case_id": 14, "diagnosis_number": 8, "diagnosis": 42, "dup_pass": 6,
    "cascade_matched_term": 26, "cascade_matched_group": 24,
    "cascade_matched_code": 12, "cascade_method": 10,
    "HISTOPATHOLOGICAL SUMMARY": 60, "FINAL COMMENT": 48,
    "verdict": 12, "confirmed_term": 28, "confirmed_group": 24, "notes": 30,
}

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CASCADE_FILL = PatternFill("solid", fgColor="EFEFEF")   # grey: read-only suggestion
_HUMAN_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow: fill these in
_AMBER_FILL = PatternFill("solid", fgColor="FFD966")     # correction needed
_UNFILLED_FILL = PatternFill("solid", fgColor="FCE4D6")  # verdict still blank
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")


def write_review_workbook(
    path: str,
    header: list[str],
    rows: list[dict],
    groups: list[str],
    terms: list[str],
    taxonomy_rows: list[tuple[str, str, str]],
) -> None:
    """Write the review workbook with dropdowns, formatting and reference sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = REVIEW_SHEET

    col_letter = {name: get_column_letter(i + 1) for i, name in enumerate(header)}
    max_row = len(rows) + 1  # +1 for header

    # --- Header row ---
    for i, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _TOP

    # --- Data rows ---
    for r, row in enumerate(rows, start=2):
        for i, name in enumerate(header, start=1):
            cell = ws.cell(row=r, column=i, value=row.get(name, ""))
            if name in _CASCADE_COLS:
                cell.fill = _CASCADE_FILL
            elif name in _HUMAN_COLS:
                cell.fill = _HUMAN_FILL
            cell.alignment = _WRAP if name in _WRAP_COLS else _TOP

    # --- Column widths; hide the dup_pass column so repeats aren't obvious ---
    for name, letter in col_letter.items():
        ws.column_dimensions[letter].width = _WIDTHS.get(name, 14)
    if "dup_pass" in col_letter:
        ws.column_dimensions[col_letter["dup_pass"]].hidden = True

    # Freeze header row + the case/diagnosis identity columns (everything left of E).
    ws.freeze_panes = "E2"

    # --- Hidden dropdown-source sheet ---
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists["A1"] = "groups"
    lists["B1"] = "terms"
    for i, g in enumerate(groups, start=2):
        lists.cell(row=i, column=1, value=g)
    for i, t in enumerate(terms, start=2):
        lists.cell(row=i, column=2, value=t)
    wb.defined_names.add(DefinedName(
        "GroupList", attr_text=f"Lists!$A$2:$A${len(groups) + 1}"))
    wb.defined_names.add(DefinedName(
        "TermList", attr_text=f"Lists!$B$2:$B${len(terms) + 1}"))

    # --- Data validations (dropdowns). errorStyle=warning so a rare off-list
    #     term can still be typed, but typos are caught. ---
    def _add_dv(formula1: str, target_col: str) -> None:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.errorStyle = "warning"
        ws.add_data_validation(dv)
        letter = col_letter[target_col]
        dv.add(f"{letter}2:{letter}{max_row}")

    _add_dv('"' + ",".join(VERDICT_CHOICES) + '"', "verdict")
    _add_dv("=GroupList", "confirmed_group")
    _add_dv("=TermList", "confirmed_term")

    # --- Conditional formatting ---
    v = col_letter["verdict"]
    ct, cg = col_letter["confirmed_term"], col_letter["confirmed_group"]
    # Light amber on the confirmed_* cells when the verdict is "wrong".
    ws.conditional_formatting.add(
        f"{ct}2:{cg}{max_row}",
        FormulaRule(formula=[f'${v}2="wrong"'], fill=_AMBER_FILL),
    )
    # Salmon on a verdict cell that is still blank.
    ws.conditional_formatting.add(
        f"{v}2:{v}{max_row}",
        FormulaRule(formula=[f"LEN(${v}2)=0"], fill=_UNFILLED_FILL),
    )

    _write_instructions(wb)
    _write_taxonomy(wb, taxonomy_rows)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    lines = [
        ("Gold annotation review — instructions", True),
        ("", False),
        ("For each row (one diagnosis), set the yellow 'verdict' column:", False),
        ("  correct    — the grey cascade suggestion is right. Nothing else to do.", False),
        ("  wrong      — a cancer label applies but the suggestion is wrong.", False),
        ("               Pick confirmed_group and confirmed_term from the dropdowns.", False),
        ("  no_cancer  — this diagnosis is not a reportable neoplasm.", False),
        ("  uncertain  — hedged / cannot be determined from the text.", False),
        ("", False),
        ("Blank verdict cells are highlighted salmon; 'wrong' rows highlight the", False),
        ("confirmed_* cells amber. Use the report context columns (HIST SUMMARY,", False),
        ("FINAL COMMENT) to judge negation and hedging.", False),
        ("The ICD-O code is filled in automatically from your group+term choice.", False),
        ("See the 'Taxonomy' sheet for the full list of valid groups and terms.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 90


def _write_taxonomy(wb: Workbook, taxonomy_rows: list[tuple[str, str, str]]) -> None:
    ws = wb.create_sheet("Taxonomy")
    for i, name in enumerate(("Group", "Term", "Code"), start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    for r, (group, term, code) in enumerate(taxonomy_rows, start=2):
        ws.cell(row=r, column=1, value=group)
        ws.cell(row=r, column=2, value=term)
        ws.cell(row=r, column=3, value=code)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"


def read_review_rows(path: str) -> list[dict]:
    """Return the Review sheet as a list of header-keyed dicts (values as str)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[REVIEW_SHEET]
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    rows: list[dict] = []
    for raw in it:
        if all(c is None for c in raw):
            continue
        rows.append({
            header[i]: ("" if i >= len(raw) or raw[i] is None else str(raw[i]))
            for i in range(len(header))
        })
    wb.close()
    return rows
